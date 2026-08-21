"""
FastAPI Backend for KBLI Code Lookup
Pattern matching + AI-Enhanced Hybrid Search
Endpoints: /lookup, /lookup/batch, /search, /search/smart, /search/hybrid
"""

import json
import re
import os
import sys
import asyncio
from uuid import uuid4
from pathlib import Path
from typing import Optional
from io import BytesIO
from contextlib import asynccontextmanager
from zipfile import BadZipFile

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from pydantic import BaseModel
from starlette.background import BackgroundTask
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.exceptions import InvalidFileException
from dotenv import load_dotenv
from openai import AsyncOpenAI

# Windows consoles often default to cp1252, which can crash on emoji/status logs.
for stream in (sys.stdout, sys.stderr):
    try:
        stream.reconfigure(encoding="utf-8")
    except Exception:
        pass

# Import Hybrid Search Engine
try:
    from backend.feedback_store import FeedbackStore
    from backend.hybrid_search import BM25, HybridSearchEngine, LocalVectorStore, reciprocal_rank_fusion
    from backend.query_understanding import (
        QueryUnderstandingService,
        build_retrieval_queries,
        local_query_understanding,
    )
except ImportError:
    from feedback_store import FeedbackStore
    from hybrid_search import BM25, HybridSearchEngine, LocalVectorStore, reciprocal_rank_fusion
    from query_understanding import QueryUnderstandingService, build_retrieval_queries, local_query_understanding

# Load environment variables
load_dotenv()
OPENAI_TIMEOUT_SECONDS = float(os.getenv("OPENAI_TIMEOUT_SECONDS", "30"))
CLASSIFICATION_MODEL = os.getenv("OPENAI_MODEL", "gpt-5.6-terra")
KBJI_RERANK_MODEL = os.getenv("KBJI_RERANK_MODEL", CLASSIFICATION_MODEL)
QUERY_UNDERSTANDING_MODEL = os.getenv("QUERY_UNDERSTANDING_MODEL", CLASSIFICATION_MODEL)
OPENAI_REASONING_EFFORT = os.getenv("OPENAI_REASONING_EFFORT", "high")
FEEDBACK_DB_PATH = Path(
    os.getenv(
        "FEEDBACK_DB_PATH",
        str(Path(__file__).parent.parent / "data" / "feedback.sqlite3"),
    )
)

# Initialize the async client once; synchronous SDK calls would block FastAPI's event loop.
async_openai_client: AsyncOpenAI = None
try:
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        async_openai_client = AsyncOpenAI(
            api_key=api_key,
            timeout=OPENAI_TIMEOUT_SECONDS,
            max_retries=2,
        )
        print("✅ OpenAI client initialized for smart search")
    else:
        print("⚠️ No OPENAI_API_KEY found - smart search disabled")
except Exception as e:
    print(f"⚠️ OpenAI initialization failed: {e}")

query_understanding_service = QueryUnderstandingService(
    async_openai_client,
    model=QUERY_UNDERSTANDING_MODEL,
)
feedback_store = FeedbackStore(FEEDBACK_DB_PATH)

# Global Hybrid Search Engine
hybrid_search_engine: HybridSearchEngine = None

# Global lookup dictionary: kode -> info
kbli_lookup: dict[str, dict] = {}
kbli_raw_data: list[dict] = []  # Raw data for hybrid search
kbji_lookup: dict[str, dict] = {}
kbji_raw_data: list[dict] = []
kbji_bm25: BM25 = None
kbji_vector_store: LocalVectorStore = None
kbji_hybrid_ready = False

# Search aliases connect common Indonesian job names to the task-based wording
# used by KBJI. They are indexed, but never replace the official title.
KBJI_CODE_ALIASES = {
    "3511.01": "operator komputer operator sistem komputer",
    "4110": "administrasi kantor tenaga administrasi tata usaha",
    "4110.00": "admin sekolah administrasi sekolah tata usaha sekolah tenaga administrasi sekolah",
    "4132": "entri data input data operator data",
    "4132.01": "operator sekolah operator dapodik operator data sekolah entri data sekolah",
    "4132.02": "petugas input data sekolah admin data sekolah",
    "5414": "satpol pp polisi pamong praja ketertiban patroli pengamanan penertiban",
}


def prepare_kbji_search_entry(entry: dict) -> dict:
    """Attach clean weighted text used only by KBJI retrieval."""
    prepared = dict(entry)
    code = str(prepared.get("kode_kbji", "")).strip()
    title = str(prepared.get("judul", "")).strip()
    description = str(prepared.get("deskripsi", "")).strip()
    aliases = KBJI_CODE_ALIASES.get(code, "")
    prepared["search_aliases"] = aliases
    # Repeating the official title gives it more lexical weight than a long
    # description containing a generic context word such as "sekolah".
    prepared["search_text"] = f"{title} {title} {title} {aliases} {description}".strip()
    return prepared

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Load KBLI data into memory and initialize Hybrid Search Engine"""
    global kbli_lookup, kbli_raw_data, kbji_lookup, kbji_raw_data, kbji_bm25, kbji_vector_store, kbji_hybrid_ready, hybrid_search_engine, async_openai_client
    
    json_path = Path(__file__).parent.parent / "kbli_parsed_fast.json"
    if not json_path.exists():
        print("ERROR: kbli_parsed_fast.json not found!")
        # We still yield because FastAPI expects the lifespan to yield even on failure
        yield
        return
    
    with open(json_path, 'r', encoding='utf-8') as f:
        kbli_data = json.load(f)
    
    # Store raw data for hybrid search
    kbli_raw_data = kbli_data
    
    # Build lookup dictionary - normalize keys
    for entry in kbli_data:
        code = entry.get("kode_kbli", "").strip()
        if code:
            # Store both original and zero-padded versions
            kbli_lookup[code] = {
                "kode": code,
                "judul": entry.get("judul", ""),
                "hierarki": entry.get("hierarki", ""),
                "cakupan": entry.get("cakupan", "")[:500],  # Truncate cakupan
                "metadata": entry.get("metadata", {})
            }
            # Also store padded version for 5-digit lookup
            if len(code) < 5 and code.isdigit():
                padded = code.zfill(5)
                kbli_lookup[padded] = kbli_lookup[code]
    
    print(f"✅ Loaded {len(kbli_lookup)} KBLI entries into lookup dictionary")

    kbji_path = Path(__file__).parent.parent / "kbji_parsed.json"
    if kbji_path.exists():
        with open(kbji_path, 'r', encoding='utf-8') as f:
            kbji_raw_data = json.load(f)
        import re
        kbji_regex = re.compile(r'^\d{1,4}(\.\d{2})?$')
        valid_kbji = []

        for raw_entry in kbji_raw_data:
            entry = prepare_kbji_search_entry(raw_entry)
            code = str(entry.get("kode_kbji", "")).strip()
            judul = str(entry.get("judul", "")).strip()
            if code and code not in kbji_lookup and len(judul) >= 4 and kbji_regex.match(code):
                kbji_lookup[code] = entry
                valid_kbji.append(entry)
                
        # Replace the raw data with cleaned data
        kbji_raw_data.clear()
        kbji_raw_data.extend(valid_kbji)

        print(f"✅ Loaded {len(kbji_lookup)} KBJI entries into lookup dictionary")
    else:
        print("⚠️ kbji_parsed.json not found - KBJI search disabled")
    
    # Initialize Hybrid Search Engine
    api_key = os.getenv("OPENAI_API_KEY")
    if api_key:
        try:
            if async_openai_client is None:
                async_openai_client = AsyncOpenAI(
                    api_key=api_key,
                    timeout=OPENAI_TIMEOUT_SECONDS,
                    max_retries=2,
                )
            query_understanding_service.client = async_openai_client
            hybrid_search_engine = HybridSearchEngine(async_openai_client)
            
            # Initialize with embeddings cache in parent directory
            cache_dir = Path(__file__).parent.parent
            await hybrid_search_engine.initialize(kbli_raw_data, cache_dir=cache_dir)
            
            print("✅ Hybrid Search Engine initialized!")

            if kbji_raw_data:
                print("🔨 Building KBJI BM25 index...")
                kbji_bm25 = BM25()
                kbji_bm25.fit(kbji_raw_data, text_fields=["search_text"])

                print("🔨 Building KBJI Vector Store...")
                kbji_vector_store = LocalVectorStore(
                    async_openai_client,
                    cache_file="kbji_embeddings_cache.pkl"
                )
                await kbji_vector_store.build_index(
                    kbji_raw_data,
                    text_fields=["judul", "search_aliases", "deskripsi"],
                    cache_dir=cache_dir
                )
                kbji_hybrid_ready = True
                print("✅ KBJI Hybrid Search Engine initialized!")
        except Exception as e:
            print(f"⚠️ Hybrid Search initialization failed: {e}")
            hybrid_search_engine = None
            kbji_bm25 = None
            kbji_vector_store = None
            kbji_hybrid_ready = False
    else:
        print("⚠️ No OPENAI_API_KEY - Hybrid Search disabled")

    yield
    # Shutdown logic (none needed here but this is where it would go)
    print("Shutting down...")

app = FastAPI(
    title="KBLI 2025 Code Lookup",
    description="Pattern-matching + AI-Enhanced Hybrid Semantic Search for KBLI codes",
    version="3.0.0",  # Major version bump for Hybrid Search
    lifespan=lifespan
)

# CORS for frontend
default_origins = (
    "https://kbli2025.alwansegeramutasi.my.id,"
    "http://localhost:3001,http://127.0.0.1:3001"
)
allowed_origins = [
    origin.strip()
    for origin in os.getenv("CORS_ALLOW_ORIGINS", default_origins).split(",")
    if origin.strip()
]
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=False,
    allow_methods=["GET", "POST"],
    allow_headers=["Accept", "Content-Type"],
)

MAX_UPLOAD_BYTES = int(os.getenv("MAX_UPLOAD_BYTES", str(10 * 1024 * 1024)))


async def read_xlsx_upload(file: UploadFile) -> bytes:
    """Validate an OOXML workbook before loading it fully into memory."""
    filename = file.filename or ""
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are supported")

    content = await file.read(MAX_UPLOAD_BYTES + 1)
    if len(content) > MAX_UPLOAD_BYTES:
        raise HTTPException(
            status_code=413,
            detail=f"File exceeds the {MAX_UPLOAD_BYTES // (1024 * 1024)} MB upload limit",
        )
    if not content:
        raise HTTPException(status_code=400, detail="Uploaded workbook is empty")
    return content


def load_xlsx(content: bytes, *, read_only: bool = False):
    """Convert malformed workbook errors into a stable client-facing response."""
    try:
        return openpyxl.load_workbook(BytesIO(content), read_only=read_only)
    except (InvalidFileException, BadZipFile, OSError, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Invalid or corrupted .xlsx workbook") from exc


def safe_output_stem(filename: str | None) -> str:
    """Return a filesystem-safe, bounded stem for a user supplied filename."""
    basename = str(filename or "workbook").replace("\\", "/").rsplit("/", 1)[-1]
    stem = basename.rsplit(".", 1)[0]
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    return (cleaned or "workbook")[:80]

def extract_kbli_codes(text: str) -> list[str]:
    """Extract potential KBLI codes from text using regex"""
    if not text:
        return []
    
    text = str(text).strip()
    
    # Pattern: 5-digit numbers (standard KBLI)
    pattern_5digit = r'\b(\d{5})\b'
    codes = re.findall(pattern_5digit, text)
    
    # Also try 2-4 digit if nothing found (might be category/golongan)
    if not codes:
        pattern_short = r'\b(\d{2,4})\b'
        codes = re.findall(pattern_short, text)
    
    return list(dict.fromkeys(codes))  # Remove duplicates, preserve order

def lookup_code(code: str) -> dict:
    """Lookup a single KBLI code"""
    code = str(code).strip()
    
    # Try exact match first
    if code in kbli_lookup:
        return {**kbli_lookup[code], "status": "found"}
    
    # Try zero-padded version
    if code.isdigit():
        padded = code.zfill(5)
        if padded in kbli_lookup:
            return {**kbli_lookup[padded], "status": "found"}
    
    # Not found
    return {
        "kode": code,
        "judul": "",
        "hierarki": "",
        "cakupan": "",
        "metadata": {},
        "status": "not_found"
    }


def format_code_matches(codes: list[str]) -> tuple[list[str], list[str], list[str]]:
    """Format titles and hierarchies without leaking state between codes."""
    titles = []
    hierarchies = []
    found_codes = []
    for code in codes:
        result = lookup_code(code)
        if result["status"] == "found":
            normalized_code = result["kode"]
            found_codes.append(normalized_code)
            titles.append(f"[{normalized_code}] {result['judul']}")
            hierarchies.append(f"[{normalized_code}] {result['hierarki']}")
        else:
            titles.append(f"[{code}] Not Found")
            hierarchies.append(f"[{code}] -")
    return titles, hierarchies, found_codes

def _tokenize_text(text: str) -> list[str]:
    """Tokenize Indonesian text for transparent local matching reasons."""
    if not text:
        return []
    return re.findall(r"\b[a-z0-9]{3,}\b", str(text).lower())


def _feedback_terms(understanding: dict | None) -> list[str]:
    """Collect stable intent terms without persisting the full LLM exchange."""
    if not understanding:
        return []
    terms = understanding.get("core_terms", []) + understanding.get("context_terms", [])
    return [str(term)[:80] for term in terms if str(term).strip()][:30]


def attach_feedback_session(
    taxonomy: str,
    query: str,
    results: list[dict],
    method: str,
    client_id: str | None = None,
    understanding: dict | None = None,
) -> tuple[list[dict], str, dict]:
    """Apply prior judgments, then register exactly what the user was shown."""
    terms = _feedback_terms(understanding)
    ranked_results, learning = feedback_store.apply_feedback(
        taxonomy,
        query,
        results,
        client_id=client_id,
        terms=terms,
    )
    candidate_codes = [
        str(
            result.get("code")
            or result.get("kode_kbli")
            or result.get("kode_kbji")
            or ""
        )
        for result in ranked_results
    ]
    session_id = feedback_store.create_impression(
        taxonomy,
        query,
        candidate_codes,
        method,
        terms=terms,
    )
    return ranked_results, session_id, learning

LOCAL_QUERY_EXPANSIONS = [
    (("satpol pp", "polisi pamong praja", "satuan polisi pamong praja"), ["84119", "administrasi pemerintahan", "ketertiban umum", "penegakan perda"]),
    (("warung makan", "rumah makan", "restoran", "jualan makanan"), ["56101", "rumah makan", "restoran", "penyediaan makanan"]),
    (("warung tenda", "food truck", "gerobak makanan", "makanan keliling"), ["56102", "warung tenda", "makanan keliling", "bangunan tidak tetap"]),
    (("bengkel motor", "servis motor", "service motor"), ["95320", "reparasi sepeda motor", "perawatan sepeda motor"]),
    (("ojek", "gojek", "grab bike"), ["49296", "angkutan ojek", "sepeda motor"]),
    (("pangkas rambut", "potong rambut", "barbershop", "salon"), ["96210", "pangkas rambut", "penataan rambut"]),
    (("jualan pulsa", "konter pulsa", "pulsa"), ["61105", "jasa sistem komunikasi", "telekomunikasi"]),
    (("warung kelontong", "toko kelontong", "warung madura"), ["47111", "perdagangan eceran", "berbagai macam barang", "kelontong"]),
    (
        (
            "es teler", "es doger", "es cincau", "es buah", "jus buah",
            "jualan minuman", "penjual minuman", "minuman racikan",
        ),
        [
            "56306", "56304", "56303", "penyediaan minuman",
            "minuman siap dikonsumsi", "proses pembuatan", "kedai minuman",
            "keliling", "tempat tidak tetap",
        ],
    ),
]

KBJI_SCHOOL_DATA_TRIGGERS = (
    "operator sekolah",
    "operator dapodik",
    "admin sekolah",
    "administrasi sekolah",
    "dapodik",
    "data sekolah",
    "data siswa",
    "memutakhirkan data",
    "memasukkan data siswa",
)

KBJI_QUERY_EXPANSIONS = [
    (
        ("satpol pp", "polisi pamong praja", "satuan polisi pamong praja"),
        ["5414", "penjaga keamanan", "petugas patroli keamanan", "ketertiban", "patroli", "penertiban", "pengamanan"],
    ),
    (
        KBJI_SCHOOL_DATA_TRIGGERS,
        [
            "4132.01",
            "4132.02",
            "4110.00",
            "operator entri data",
            "petugas input data",
            "tenaga perkantoran umum",
            "data sekolah",
            "administrasi sekolah",
        ],
    ),
]

def expand_local_keywords(query: str) -> list[str]:
    """Expand common informal business terms into KBLI-friendly keywords."""
    query_lower = str(query).lower()
    keywords = _tokenize_text(query)

    for triggers, expansions in LOCAL_QUERY_EXPANSIONS:
        if any(trigger in query_lower for trigger in triggers):
            for keyword in expansions:
                if keyword not in keywords:
                    keywords.append(keyword)

    return keywords

def expand_kbji_keywords(query: str) -> list[str]:
    """Expand informal job terms into KBJI-friendly occupation keywords."""
    query_lower = str(query).lower()
    keywords = list(dict.fromkeys(_tokenize_text(query)))

    for triggers, expansions in KBJI_QUERY_EXPANSIONS:
        if any(trigger in query_lower for trigger in triggers):
            for keyword in expansions:
                if keyword not in keywords:
                    keywords.append(keyword)

    return keywords

def get_manual_classifications(query: str) -> list[dict]:
    """Curated classifications for common public-sector/job-title inputs."""
    query_lower = str(query).lower()

    if any(term in query_lower for term in ("satpol pp", "polisi pamong praja", "satuan polisi pamong praja")):
        kbji_match = lookup_kbji_code("5414")
        return [{
            "code": "84119",
            "judul": "KEGIATAN PEMERINTAHAN LAINNYA",
            "hierarki": "P ADMINISTRASI PEMERINTAHAN DAN PERTAHANAN, SERTA JAMINAN > 84 ADMINISTRASI PEMERINTAHAN",
            "cakupan": (
                "Kegiatan lembaga/badan/instansi pemerintah lainnya dengan tugas khusus serta "
                "kesekretariatannya yang belum tercakup dalam kelompok 84111 s.d. 84115."
            ),
            "score": 10000,
            "matched_keywords": ["satpol pp", "polisi pamong praja", "ketertiban umum", "penegakan perda"],
            "reasoning": (
                "Satpol PP diklasifikasikan ke KBLI 84119 karena KBLI mengklasifikasikan "
                "kegiatan instansi/lapangan usaha, bukan jabatan orang. Satpol PP adalah perangkat "
                "pemerintah daerah dengan tugas khusus penegakan perda, ketertiban umum, dan "
                "ketenteraman masyarakat. Kegiatan ini bukan usaha keamanan swasta, bukan lembaga "
                "legislatif, bukan lembaga keuangan/perpajakan, dan bukan lembaga perencanaan, "
                "sehingga paling tepat masuk kegiatan pemerintahan lainnya."
            ),
            "kbji": {
                "code": kbji_match.get("kode_kbji", "5414"),
                "judul": kbji_match.get("judul", "Penjaga Keamanan"),
                "deskripsi": kbji_match.get("deskripsi", ""),
                "reasoning": (
                    "KBJI mengklasifikasikan pekerjaan/jabatan orangnya. Tugas lapangan Satpol PP "
                    "dekat dengan pekerjaan pengamanan, patroli, penjagaan ketertiban, dan "
                    "penertiban. Di data KBJI lokal, 5414 Penjaga Keamanan memuat tugas menjaga "
                    "keamanan bangunan/wilayah, patroli, pengendalian akses, menjaga ketertiban, "
                    "dan menanggapi gangguan. Karena entri eksplisit Polisi Pamong Praja tidak "
                    "ditemukan sebagai kode rinci di data KBJI lokal, 5414 dipakai sebagai padanan "
                    "pekerjaan terdekat. Untuk konteks ASN, Polisi Pamong Praja juga memiliki "
                    "nomenklatur jabatan fungsional tersendiri, tetapi itu bukan kode KBJI."
                )
            }
        }]

    return []

def lookup_kbji_code(code: str) -> dict:
    """Lookup a KBJI code from parsed KBJI data."""
    code = str(code).strip()
    return kbji_lookup.get(code, {
        "kode_kbji": code,
        "judul": "",
        "deskripsi": "",
        "source_page": None,
        "level": "",
    })

def get_manual_kbji_classifications(query: str) -> list[dict]:
    """Curated KBJI results for terms that are not explicit KBJI titles."""
    query_lower = str(query).lower()

    def result(code: str, score: int, keywords: list[str], reasoning: str) -> dict:
        entry = lookup_kbji_code(code)
        return {
            "code": entry.get("kode_kbji", code),
            "kode_kbji": entry.get("kode_kbji", code),
            "judul": entry.get("judul", ""),
            "deskripsi": entry.get("deskripsi", "")[:500],
            "source_page": entry.get("source_page"),
            "level": entry.get("level", "rinci"),
            "score": score,
            "matched_keywords": keywords,
            "reasoning": reasoning,
        }

    if any(term in query_lower for term in ("satpol pp", "polisi pamong praja", "satuan polisi pamong praja")):
        return [result(
            "5414",
            98,
            ["satpol pp", "ketertiban", "patroli", "penertiban", "pengamanan"],
            (
                "Satpol PP bukan POLRI/TNI, sehingga tidak tepat dinaikkan ke Bintara POLRI. "
                "Untuk KBJI, yang diklasifikasikan adalah pekerjaan/jabatan orangnya. Tugas "
                "lapangan Satpol PP paling dekat dengan fungsi menjaga ketertiban, patroli, "
                "pengamanan, dan penertiban. Dalam data KBJI lokal, kelompok terdekat adalah "
                "5414 Penjaga Keamanan."
            ),
        )]

    if any(term in query_lower for term in KBJI_SCHOOL_DATA_TRIGGERS):
        return [
            result(
                "4132.01",
                98,
                ["operator", "entri data", "data sekolah"],
                (
                    "KBJI mengklasifikasikan tugas orangnya, sedangkan sekolah adalah tempat "
                    "kerjanya. Operator sekolah umumnya memasukkan, memutakhirkan, memeriksa, "
                    "dan mengirim data sekolah melalui sistem seperti Dapodik. Tugas tersebut "
                    "paling dekat dengan KBJI 4132.01 Operator Entri Data."
                ),
            ),
            result(
                "4132.02",
                93,
                ["input data", "data sekolah"],
                (
                    "KBJI 4132.02 Petugas Input Data merupakan alternatif apabila pekerjaan "
                    "lebih banyak berupa pencatatan dan pemeriksaan data sekolah, bukan "
                    "pengelolaan administrasi kantor secara luas."
                ),
            ),
            result(
                "4110.00",
                88,
                ["administrasi sekolah", "tata usaha"],
                (
                    "KBJI 4110.00 Tenaga Perkantoran Umum lebih tepat apabila operator sekolah "
                    "juga menangani surat, arsip, laporan, dan administrasi tata usaha. Pilihan "
                    "akhir bergantung pada tugas yang paling dominan."
                ),
            ),
        ]

    return []

def merge_manual_kbji_results(query: str, results: list[dict], limit: int) -> list[dict]:
    manual_results = get_manual_kbji_classifications(query)
    if not manual_results:
        return results[:limit]

    merged = []
    seen_codes = set()
    for result in manual_results + results:
        code = result.get("kode_kbji") or result.get("code")
        if code in seen_codes:
            continue
        seen_codes.add(code)
        merged.append(result)

    return merged[:limit]

def build_kbji_reasoning(query: str, result: dict, matched_keywords: list[str] | None = None) -> str:
    title = result.get("judul", "")
    code = result.get("kode_kbji", "")
    description = result.get("deskripsi", "")
    searchable = " ".join([
        str(title),
        str(description),
        str(result.get("search_aliases", "")),
    ]).lower()
    keywords = matched_keywords if matched_keywords is not None else [
        token for token in _tokenize_text(query) if token in searchable
    ]
    if keywords:
        evidence = f"kecocokan istilah {', '.join(keywords[:5])}"
    else:
        evidence = "kemiripan semantik dengan uraian tugas; kecocokan tugas tetap perlu diperiksa"
    return (
        f"Diklasifikasikan ke KBJI {code} - {title} karena KBJI mengklasifikasikan "
        f"pekerjaan/jabatan orangnya dan kandidat ditemukan berdasarkan {evidence}. "
        f"Ringkasan cakupan KBJI: {description[:260]}"
    )

def search_kbji_entries(
    query: str,
    limit: int = 5,
    original_query: str | None = None,
) -> list[dict]:
    """Local KBJI keyword search over parsed PDF data."""
    if not query or not kbji_raw_data:
        return []

    expanded_keywords = expand_kbji_keywords(query)
    filter_query = original_query or query
    query_terms = set(local_query_understanding(filter_query, "kbji")["core_terms"])
    context_only_terms = {
        "kantor", "pabrik", "pendidikan", "rumah", "sekolah", "toko", "universitas"
    }
    role_terms = query_terms - context_only_terms
    requested_context = query_terms & context_only_terms
    results = []

    for entry in kbji_raw_data:
        code = str(entry.get("kode_kbji", ""))
        title = str(entry.get("judul", ""))
        description = str(entry.get("deskripsi", ""))
        aliases = str(entry.get("search_aliases", ""))
        searchable_title = title.lower()
        searchable_description = description.lower()
        searchable_aliases = aliases.lower()

        score = 0
        matched_keywords = []

        for keyword in expanded_keywords:
            kw = str(keyword).lower().strip()
            if not kw:
                continue

            if kw == code:
                score += 5000
                matched_keywords.append(keyword)
            elif code.startswith(kw) and kw.isdigit():
                score += 2500
                matched_keywords.append(keyword)
            elif kw in searchable_aliases:
                score += 1400
                matched_keywords.append(keyword)
            elif kw in searchable_title:
                score += 800
                matched_keywords.append(keyword)
            elif kw in searchable_description:
                score += 120
                matched_keywords.append(keyword)

        query_lower = str(query).lower().strip()
        if query_lower and query_lower == searchable_title:
            score += 1200
        elif query_lower and query_lower in searchable_title:
            score += 500

        if score > 0 and entry.get("level") == "subgolongan":
            score += 50

        if score > 0:
            matched_terms = {
                token
                for keyword in matched_keywords
                for token in _tokenize_text(keyword)
            }
            # For a multi-concept query, a document matching only the workplace
            # context (for example "sekolah") is not an occupation match.
            if role_terms and not (matched_terms & role_terms):
                continue
            if requested_context and not (matched_terms & requested_context):
                continue
            matched_query_terms = matched_terms & query_terms
            required_matches = (
                len(query_terms)
                if len(query_terms) <= 3
                else max(2, (len(query_terms) + 1) // 2)
            )
            if query_terms and len(matched_query_terms) < required_matches:
                continue
            result = {
                "code": code,
                "kode_kbji": code,
                "judul": title,
                "deskripsi": description[:500],
                "source_page": entry.get("source_page"),
                "level": entry.get("level", ""),
                "score": score,
                "matched_keywords": list(dict.fromkeys(matched_keywords)),
            }
            result["reasoning"] = build_kbji_reasoning(query, entry, result["matched_keywords"])
            results.append(result)

    results.sort(key=lambda x: x["score"], reverse=True)
    if results:
        minimum_score = max(300, results[0]["score"] * 0.35)
        results = [result for result in results if result["score"] >= minimum_score]
    return merge_manual_kbji_results(original_query or query, results, limit)

async def search_kbji_hybrid_candidates(
    query: str,
    retrieval_top_k: int = 20,
    retrieval_query: str | None = None,
    semantic_query: str | None = None,
) -> dict:
    """Retrieve KBJI candidates with BM25 keyword search + semantic vector search."""
    if not kbji_hybrid_ready or not kbji_bm25 or not kbji_vector_store:
        return {
            "results": search_kbji_entries(query, retrieval_top_k),
            "bm25_top": 0,
            "vector_top": 0,
            "total_candidates": 0,
        }

    local_expansion = " ".join(str(keyword) for keyword in expand_kbji_keywords(query))
    expanded_query = f"{retrieval_query or query} {local_expansion}".strip()
    bm25_task = asyncio.create_task(
        asyncio.to_thread(kbji_bm25.search, expanded_query, retrieval_top_k)
    )
    vector_task = asyncio.create_task(
        kbji_vector_store.search(semantic_query or query, retrieval_top_k)
    )
    bm25_results, vector_results = await asyncio.gather(bm25_task, vector_task)

    fused_ranking = reciprocal_rank_fusion([bm25_results, vector_results], k=60)
    candidates = []
    seen_codes = set()

    for doc_idx, rrf_score in fused_ranking[:retrieval_top_k]:
        entry = kbji_raw_data[doc_idx]
        code = str(entry.get("kode_kbji", ""))
        if code in seen_codes:
            continue
        seen_codes.add(code)
        candidates.append({
            "code": code,
            "kode_kbji": code,
            "judul": entry.get("judul", ""),
            "deskripsi": entry.get("deskripsi", "")[:500],
            "source_page": entry.get("source_page"),
            "level": entry.get("level", ""),
            "score": rrf_score * 1000,
            "rrf_score": rrf_score,
            "matched_keywords": [],
            "reasoning": build_kbji_reasoning(query, entry),
        })

    return {
        "results": merge_manual_kbji_results(query, candidates, retrieval_top_k),
        "bm25_top": len(bm25_results),
        "vector_top": len(vector_results),
        "total_candidates": len(fused_ranking),
    }

def build_local_reasoning(query: str, result: dict, matched_keywords: list[str] | None = None) -> str:
    """
    Build a human-readable classification reason without requiring an LLM.
    This keeps the UI explainable when OPENAI_API_KEY is not configured.
    """
    title = result.get("judul", "")
    hierarchy = result.get("hierarki", "")
    scope = result.get("cakupan", "")
    code = result.get("code") or result.get("kode") or result.get("kode_kbli", "")

    query_lower = str(query).lower()
    fresh_drink_terms = (
        "es teler", "es doger", "es cincau", "es buah", "jus buah",
        "jualan minuman", "penjual minuman", "minuman racikan",
    )
    is_fresh_drink = (
        any(term in query_lower for term in fresh_drink_terms)
        or "minuman siap dikonsumsi" in query_lower
        or ("teler" in query_lower and "minuman" in query_lower)
    )
    if is_fresh_drink:
        contextual_reasons = {
            "56306": (
                "Input menunjukkan minuman yang diracik atau dibuat untuk langsung dikonsumsi, "
                "bukan sekadar penjualan kembali minuman kemasan. KBLI 56306 paling dekat bila "
                "penjualannya keliling atau memakai tempat tidak tetap; cakupannya juga memberi "
                "contoh minuman es sejenis seperti es doger dan es cincau. Jika usaha menetap di "
                "bangunan permanen, 56303 lebih tepat; jika berupa kedai/tenda bongkar-pasang, "
                "pertimbangkan 56304."
            ),
            "56304": (
                "Input menunjukkan penyajian minuman racikan siap konsumsi. KBLI 56304 sesuai "
                "apabila usaha dijalankan sebagai kedai atau tenda bongkar-pasang, seperti kedai "
                "jus. Bila berjualan keliling/tempat tidak tetap gunakan 56306, sedangkan bangunan "
                "permanen lebih dekat ke 56303."
            ),
            "56303": (
                "Input menunjukkan penyediaan minuman siap konsumsi. KBLI 56303 sesuai apabila "
                "es teler disajikan dari rumah minum atau kafe di bangunan permanen. Bila model "
                "usahanya kedai/tenda atau berkeliling, 56304 atau 56306 lebih tepat."
            ),
            "47222": (
                "KBLI 47222 hanya tepat bila kegiatan utamanya menjual kembali minuman "
                "nonalkohol dan tidak untuk langsung dikonsumsi di tempat. Karena es teler "
                "umumnya diracik untuk langsung diminum, kode penyediaan minuman 56303-56306 "
                "biasanya lebih kuat kecuali pengguna menjelaskan bahwa produknya hanya dijual "
                "dalam kemasan."
            ),
        }
        if code in contextual_reasons:
            return f"Diklasifikasikan ke {code} - {title}. {contextual_reasons[code]}"

    searchable_title = title.lower()
    searchable_hierarchy = hierarchy.lower()
    searchable_scope = scope.lower()

    keywords = []
    for kw in matched_keywords or _tokenize_text(query):
        clean = str(kw).strip().lower()
        if clean and clean not in keywords:
            keywords.append(clean)

    title_matches = [kw for kw in keywords if kw in searchable_title]
    hierarchy_matches = [kw for kw in keywords if kw in searchable_hierarchy and kw not in title_matches]
    scope_matches = [
        kw for kw in keywords
        if kw in searchable_scope and kw not in title_matches and kw not in hierarchy_matches
    ]

    reason_parts = []
    if title_matches:
        reason_parts.append(f"judul KBLI memuat kata kunci {', '.join(title_matches[:4])}")
    if hierarchy_matches:
        reason_parts.append(f"hierarki/sektor terkait dengan {', '.join(hierarchy_matches[:3])}")
    if scope_matches:
        reason_parts.append(f"cakupan menjelaskan aktivitas yang terkait dengan {', '.join(scope_matches[:3])}")

    if not reason_parts:
        reason_parts.append("kode ini muncul sebagai kandidat terdekat berdasarkan kemiripan kata pada judul, hierarki, dan cakupan KBLI")

    return (
        f"Diklasifikasikan ke {code} - {title} karena "
        f"{'; '.join(reason_parts)}. Tetap verifikasi aktivitas utama usaha jika input memuat lebih dari satu pekerjaan."
    )

class LookupRequest(BaseModel):
    code: str

class LookupResponse(BaseModel):
    code: str
    found: bool
    judul: str
    hierarki: str
    cakupan: str


class FeedbackRequest(BaseModel):
    session_id: str
    client_id: str
    selected_code: Optional[str] = None
    no_match: bool = False

@app.get("/")
async def root():
    return {
        "status": "ok",
        "service": "KBLI 2025 Code Lookup v2.0",
        "total_entries": len(kbli_lookup)
    }

@app.get("/health")
async def health():
    return {
        "status": "healthy",
        "entries_loaded": len(kbli_lookup),
        "kbji_entries_loaded": len(kbji_lookup),
        "method": "pattern_matching",
        "kbli_hybrid_ready": bool(hybrid_search_engine and hybrid_search_engine.is_ready),
        "kbji_hybrid_ready": kbji_hybrid_ready
    }


@app.post("/feedback")
async def submit_feedback(request: FeedbackRequest):
    """Record an explicit result selection or a no-match judgment."""
    if len(request.session_id) > 64 or len(request.client_id) > 80:
        raise HTTPException(status_code=400, detail="Invalid feedback payload")
    if request.selected_code and len(request.selected_code) > 20:
        raise HTTPException(status_code=400, detail="Invalid classification code")

    try:
        return feedback_store.save_feedback(
            request.session_id,
            request.client_id,
            selected_code=request.selected_code,
            no_match=request.no_match,
        )
    except LookupError as error:
        raise HTTPException(status_code=404, detail=str(error)) from error
    except ValueError as error:
        raise HTTPException(status_code=400, detail=str(error)) from error


@app.get("/feedback/stats")
async def feedback_stats():
    """Expose aggregate counts only; raw queries and client IDs stay private."""
    return feedback_store.stats()

@app.get("/stats")
async def stats():
    """Get statistics about loaded KBLI data"""
    return {
        "total_entries": len(kbli_lookup),
        "total_kbji_entries": len(kbji_lookup),
        "sample_codes": list(kbli_lookup.keys())[:10]
    }

@app.get("/search")
async def search_kbli(
    q: str = Query(min_length=2, max_length=200),
    limit: int = Query(default=10, ge=1, le=50),
):
    """
    Search KBLI by keyword in title, hierarchy, or description.
    Supports fuzzy matching.
    """
    if not q or len(q) < 2:
        return {"results": [], "query": q, "total": 0}
    
    q_lower = q.lower()
    results = []
    
    for code, info in kbli_lookup.items():
        # Search in title, hierarchy, and cakupan
        searchable = f"{info['judul']} {info['hierarki']} {info.get('cakupan', '')}".lower()
        
        # Simple relevance scoring
        score = 0
        if q_lower in searchable:
            # Exact substring match
            score = 100
            # Bonus if in title
            if q_lower in info['judul'].lower():
                score += 50
            # Bonus if at start
            if searchable.startswith(q_lower):
                score += 25
        else:
            # Fuzzy match - check if all query words are present
            query_words = q_lower.split()
            matches = sum(1 for word in query_words if word in searchable)
            if matches > 0:
                score = (matches / len(query_words)) * 50
        
        if score > 0:
            results.append({
                "code": code,
                "judul": info["judul"],
                "hierarki": info["hierarki"],
                "score": score
            })
    
    # Sort by score descending
    results.sort(key=lambda x: x["score"], reverse=True)
    
    return {
        "results": results[:limit],
        "query": q,
        "total": len(results)
    }

@app.get("/autocomplete")
async def autocomplete(
    q: str = Query(min_length=1, max_length=100),
    limit: int = Query(default=5, ge=1, le=20),
):
    """
    Autocomplete suggestions for KBLI codes and titles.
    Returns quick suggestions as user types.
    """
    if not q or len(q) < 1:
        return {"suggestions": []}
    
    q_lower = q.lower()
    suggestions = []
    
    for code, info in kbli_lookup.items():
        # Match by code prefix
        if code.startswith(q):
            suggestions.append({
                "type": "code",
                "code": code,
                "judul": info["judul"],
                "match": f"{code} - {info['judul'][:60]}..."
            })
        # Match by title prefix
        elif info['judul'].lower().startswith(q_lower):
            suggestions.append({
                "type": "title",
                "code": code,
                "judul": info["judul"],
                "match": f"{info['judul'][:60]}... ({code})"
            })
        # Match by word in title
        elif any(word.startswith(q_lower) for word in info['judul'].lower().split()):
            suggestions.append({
                "type": "word",
                "code": code,
                "judul": info["judul"],
                "match": f"{info['judul'][:60]}... ({code})"
            })
        
        if len(suggestions) >= limit * 3:  # Get more for sorting
            break
    
    # Prioritize: code matches > title prefix > word matches
    suggestions.sort(key=lambda x: (
        0 if x["type"] == "code" else 1 if x["type"] == "title" else 2,
        x["code"]
    ))
    
    return {"suggestions": suggestions[:limit]}

async def expand_query_with_ai(query: str) -> dict:
    """
    Use OpenAI to expand informal Indonesian query into KBLI terminology.
    Returns expanded keywords for better search matching.
    """
    if not async_openai_client:
        keywords = expand_local_keywords(query)
        return {"expanded": query, "keywords": keywords or [query], "ai_used": False}

    system_prompt = """ROLE: Anda adalah Ahli Klasifikasi Statistik BPS (Badan Pusat Statistik) khusus KBLI 2025.
TUGAS: Terjemahkan query informal user menjadi KATA KUNCI TEKNIS KBLI 2025 yang presisi.

PRINSIP DASAR KBLI (Metode Top-Down & Cakupan):
1. Tentukan Aktivitas Utama (Principal Activity) berdasarkan Nilai Tambah terbesar (sumber penghasilan utama).
2. Bedakan Jelas:
   - PERDAGANGAN (Kat G): Hanya jual beli tanpa merubah bentuk.
   - INDUSTRI (Kat C): Ada proses perubahan bentuk/fisik/kimiawi. Jika outsourcing total & punya bahan baku -> INDUSTRI. Jika tidak punya bahan baku -> PERDAGANGAN.
   - PERTANIAN (Kat A): Budidaya alam (tanam, ternak).

ATURAN KHUSUS (WAJIB PATUH):
1. HAPUS PELAKU: "tukang", "penjual", "pembuat", "pengusaha", "juragan", "ahli", "teknisi". Fokus pada KEGIATAN (misal: "memasak", "menjual") atau OBJEK (misal: "nasi goreng", "baju", "rambut").
2. HAPUS KATA UMUM/IRRELEVAN: "jasa", "usaha", "bisnis", "wanita", "pria", "sukses", "kegiatan", "aktivitas" (kecuali spesifik seperti 'jasa keuangan' atau 'aktivitas profesional').
3. PERDAGANGAN ECERAN (Kategori 47):
   - Jika ada kata "ONLINE", "INTERNET", "E-COMMERCE", "SHOPEE", "TOKOPEDIA" -> Wajib sertakan kata kunci "4791", "MELALUI POS", "INTERNET".
   - Jika ada kata "KELILING", "KAKI LIMA", "GEROBAK" -> Wajib sertakan "478", "KAKI LIMA".
   - Jika "TOKO", "BUTIK", "KIOS" atau diam -> Asumsikan toko fisik (471-477).
   - "WARUNG" / "TOKO KELONTONG" (campuran) -> "471", "BERBAGAI MACAM BARANG", "MINIMARKET".
4. INDUSTRI vs JASA:
   - "Tukang Las" -> "JASA PENGELASAN" (bukan industri mesin).
   - "Konveksi" (membuat baju) -> "INDUSTRI PAKAIAN JADI" (bukan penjahit).
   - "Permak Levis" / "Penjahit" -> "REPARASI", "PAKAIAN", "VERMAK".
   - "Bengkel Motor" -> "REPARASI", "PERAWATAN", "SEPEDA MOTOR".
5. PROFESI BUKAN LAPANGAN USAHA (PNS, ASN, PPPK, Satpol PP, dll):
   - KBLI adalah klasifikasi usaha, bukan pekerjaan.
   - Jika input adalah profesi perorangan seperti "PNS", "ASN", "PPPK", "Satpol PP", "Pegawai Negeri", wajib output: BUKAN KBLI.

FORMAT OUTPUT:
Hanya 2-6 kata kunci paling relevan, dipisahkan koma, lowercase. Urutkan dari yang paling spesifik/penting.
Jika Anda yakin 100% dengan Kode KBLI 4-5 digit yang sesuai, SERTAKAN KODE TERSEBUT di awal.

CONTOH:
Input: "Jualan baju di shopee"
Output: 4791, perdagangan eceran, melalui internet, pakaian

Input: "Warung madura jual beras rokok sabun"
Output: 4711, perdagangan eceran, berbagai macam barang, kelontong

Input: "Bikin keripik singkong di rumah sendiri"
Output: industri, makanan, keripik, singkong

Input: "Tukang pangkas rambut gaul"
Output: pangkas rambut, 9611, salon

Input: "Jasa angkut barang pindahan rumah"
Output: 494, angkutan jalan, pindahan
"""

    try:
        response = await async_openai_client.chat.completions.create(
            model="gpt-5.6-terra",
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": f"Input: \"{query}\""}
            ],
            max_completion_tokens=600,
            reasoning_effort=OPENAI_REASONING_EFFORT,
        )
        
        expanded = response.choices[0].message.content.strip()
        # Clean up output to get pure keywords
        raw_keywords = [k.strip().lower() for k in expanded.split(",")]
        # Filter empty strings and strict stop words cleanup
        stop_words = {"jasa", "usaha", "bisnis", "kegiatan", "aktivitas", "pelayanan", "tukang", "penjual", "pembuat", "ahli", "spesialis", "dan", "atau", "di", "ke", "dari", "yang"}
        keywords = [k for k in raw_keywords if k and k not in stop_words and len(k) > 1]
        
        return {
            "original": query,
            "expanded": expanded,
            "keywords": keywords,
            "ai_used": True
        }
    except Exception as e:
        print(f"OpenAI error: {e}")
        return {"expanded": query, "keywords": [query], "ai_used": False}

def search_with_keywords(keywords: list[str], limit: int = 10) -> list[dict]:
    """Search KBLI using multiple keywords with advanced scoring"""
    results = []
    has_code_hint = any(str(keyword).strip().isdigit() for keyword in keywords)
    
    for code, info in kbli_lookup.items():
        judul_lower = info['judul'].lower()
        hierarki_lower = info['hierarki'].lower()
        cakupan_lower = info.get('cakupan', '').lower()
        
        score = 0
        matched_keywords = []
        
        # Check each keyword
        for keyword in keywords:
            kw = keyword.lower().strip()
            if not kw:
                continue
            
            keyword_found = False
            
            # 0. Check for Direct Code Match (Highest Priority)
            if kw.isdigit():
                 if code == kw:
                     score += 5000 # Perfect code match
                     keyword_found = True
                     matched_keywords.append(keyword)
                 elif code.startswith(kw):
                     score += 3000 # Prefix code match (e.g. search "471" matches "47110")
                     keyword_found = True
                     matched_keywords.append(keyword)
                 continue # Skip text search if it was a digit

            # 1. Exact match in title (Highest Priority)
            if kw in judul_lower:
                # Check if it's a word boundary match (not substring)
                # Simple boundary check by splitting
                words_in_title = judul_lower.replace(",", "").replace(".", "").split()
                if kw in words_in_title:
                    score += 1500  # Huge score for exact word match
                    keyword_found = True
                    matched_keywords.append(keyword)
                elif any(kw in word for word in words_in_title):
                    score += 200   # Lower score for substring match (e.g. "jual" in "penjualan")
                    keyword_found = True
                    matched_keywords.append(keyword)
            
            # 2. Match in hierarchy (Medium priority)
            elif kw in hierarki_lower:
                score += 300
                keyword_found = True
                matched_keywords.append(keyword)
            
            # 3. Match in cakupan (Lower priority)
            elif kw in cakupan_lower:
                score += 50  # Much lower score for cakupan
                keyword_found = True
                matched_keywords.append(keyword)
        
        # Bonus: Multiple keyword matches (AND logic)
        if len(matched_keywords) > 1:
            score += len(matched_keywords) * 200
        
        # Bonus: Exact phrase match in title
        full_query = " ".join(keywords).lower()
        if full_query in judul_lower:
            score += 2000  # Huge bonus for exact phrase
        
        unique_matches = list(dict.fromkeys(matched_keywords))
        direct_code_match = any(
            str(keyword).strip().isdigit() and code.startswith(str(keyword).strip())
            for keyword in keywords
        )

        # Curated expansions contain a strong code hint. Avoid returning documents
        # that matched only one generic word such as "makan" or "barang".
        if score > 0 and (not has_code_hint or direct_code_match or len(unique_matches) >= 2):
            result = {
                "code": code,
                "judul": info["judul"],
                "hierarki": info["hierarki"],
                "cakupan": info.get("cakupan", "")[:200],
                "score": score,
                "matched_keywords": unique_matches
            }
            result["reasoning"] = build_local_reasoning(
                " ".join(keywords),
                result,
                result["matched_keywords"]
            )
            results.append(result)
    
    # Sort by score descending
    results.sort(key=lambda x: x["score"], reverse=True)
    return results[:limit]

def merge_manual_results(query: str, results: list[dict], limit: int) -> list[dict]:
    """Put curated high-confidence matches first and avoid duplicate KBLI codes."""
    manual_results = get_manual_classifications(query)
    if not manual_results:
        return results[:limit]

    merged = []
    seen_codes = set()
    for result in manual_results + results:
        code = result.get("code") or result.get("kode") or result.get("kode_kbli")
        if code in seen_codes:
            continue
        seen_codes.add(code)
        merged.append(result)

    return merged[:limit]

@app.get("/search/smart")
async def smart_search(
    q: str = Query(min_length=2, max_length=200),
    limit: int = Query(default=10, ge=1, le=20),
    client_id: Optional[str] = Query(default=None, max_length=80),
):
    """
    AI-Enhanced Smart Search.
    Uses GPT to translate informal queries into KBLI terminology.
    Falls back to pattern matching if AI unavailable.
    
    Example:
    - "tukang ojek" -> finds 49422 Angkutan Ojek
    - "warung makan" -> finds 56101 Restoran
    """
    if not q or len(q) < 2:
        return {"results": [], "query": q, "total": 0}
    
    # Step 1: Expand query with AI
    expansion = await expand_query_with_ai(q)
    
    # Step 2: Search with expanded keywords
    results = search_with_keywords(expansion["keywords"], limit)
    results = merge_manual_results(q, results, limit)
    method = "smart" if expansion.get("ai_used") else "local_keyword"
    understanding = local_query_understanding(q, "kbli")
    results, feedback_session_id, feedback_learning = attach_feedback_session(
        "kbli",
        q,
        results,
        method,
        client_id=client_id,
        understanding=understanding,
    )
    
    return {
        "query": q,
        "method": method,
        "expansion": expansion,
        "total": len(results),
        "results": results,
        "feedback_session_id": feedback_session_id,
        "feedback_learning": feedback_learning,
    }

async def rerank_kbji_candidates(
    query: str,
    candidates: list[dict],
    top_k: int = 5,
    query_context: str = "",
) -> tuple[list[dict], bool]:
    """Re-rank KBJI candidates and report whether the LLM result was used."""
    if not candidates or not async_openai_client:
        return candidates[:top_k], False
    
    candidates = candidates[:15]
    candidate_str = "\n".join([
        f"{i+1}. KODE: {c.get('kode_kbji', '')} | JUDUL: {c.get('judul', '')[:100]} | DESKRIPSI: {c.get('deskripsi', '')[:200]}"
        for i, c in enumerate(candidates)
    ])
    
    system_prompt = """Anda adalah ahli klasifikasi KBJI (Klasifikasi Baku Jabatan Indonesia).
Tugas: Evaluasi relevansi setiap kandidat KBJI terhadap query (jabatan/pekerjaan) pengguna.
ATURAN PENTING:
1. Fokus pada pekerjaan/jabatan.
2. Perhatikan konteks informal. Misal "satpol pp" relevan dengan "Polisi pamong praja" atau "kepala wilayah / ketentraman".
3. Satpol PP/Polisi Pamong Praja BUKAN POLRI/TNI. Jangan memilih Bintara POLRI, Perwira POLRI, atau jabatan TNI kecuali query eksplisit menyebut POLRI/TNI.
4. Pada frasa seperti "operator sekolah", kata "operator" adalah tugas dan "sekolah" adalah konteks tempat kerja. Utamakan entri/input data atau administrasi, bukan kepala, pengawas, atau guru sekolah.
5. Kandidat yang hanya cocok pada konteks tempat kerja tetapi tidak cocok pada tugas harus dibuang.
6. Berikan skor (0.0 - 1.0). Hanya sertakan kandidat relevan (relevance > 0.3).
7. Alasan harus 2-4 kalimat: jelaskan tugas utama, kecocokan dengan deskripsi KBJI, pembeda dari kandidat terdekat, dan konteks yang masih perlu dikonfirmasi. Jangan hanya menyebut kemiripan kata.

OUTPUT FORMAT (JSON only, no markdown):
{
  "rankings": [
    {
      "rank": 1,
      "index": <nomor kandidat 1-based>,
      "relevance": <0.0-1.0>,
      "reason": "<alasan substantif 2-4 kalimat>"
    }
  ]
}"""

    user_prompt = (
        f'Query asli: "{query}"\n'
        f'Analisis query: {query_context or "-"}\n\n'
        f'Kandidat KBJI:\n{candidate_str}\n\nOutput JSON saja.'
    )
    
    try:
        response = await async_openai_client.chat.completions.create(
            model=KBJI_RERANK_MODEL,
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt}
            ],
            reasoning_effort=OPENAI_REASONING_EFFORT,
            max_completion_tokens=1800,
            response_format={"type": "json_object"},
        )
        content = (response.choices[0].message.content or "").strip()
        if not content:
            raise ValueError("KBJI reranker returned an empty response")
        import re
        if "```" in content:
            match = re.search(r'```(?:json)?\s*(.*?)```', content, re.DOTALL)
            if match:
                content = match.group(1).strip()
        import json
        result = json.loads(content)
        rankings = result.get("rankings", [])
        
        reranked = []
        seen_indices = set()
        for r in rankings[:top_k]:
            idx = r.get("index", 1) - 1
            relevance = float(r.get("relevance", 0.0))
            if 0 <= idx < len(candidates) and idx not in seen_indices and relevance > 0.3:
                seen_indices.add(idx)
                candidate = candidates[idx].copy()
                candidate["score"] = round(relevance * 100, 2)
                candidate["reasoning"] = f"Diklasifikasikan ke KBJI {candidate['kode_kbji']} karena {r.get('reason', '')}"
                reranked.append(candidate)
        return (reranked, True) if reranked else (candidates[:top_k], False)
    except Exception as e:
        print(f"KBJI Reranking error: {e}")
        return candidates[:top_k], False

@app.get("/search/kbji")
async def kbji_search(
    q: str = Query(min_length=2, max_length=200),
    limit: int = Query(default=5, ge=1, le=20),
    client_id: Optional[str] = Query(default=None, max_length=80),
):
    """
    Search KBJI occupations using hybrid retrieval when available.
    """
    if not q or len(q) < 2:
        return {"query": q, "method": "kbji_local_keyword", "total": 0, "results": []}

    curated_results = get_manual_kbji_classifications(q)
    if curated_results:
        results = curated_results[:limit]
        understanding = local_query_understanding(q, "kbji")
        results, feedback_session_id, feedback_learning = attach_feedback_session(
            "kbji",
            q,
            results,
            "kbji_curated",
            client_id=client_id,
            understanding=understanding,
        )
        return {
            "query": q,
            "method": "kbji_curated",
            "rerank_status": "curated",
            "bm25_results": 0,
            "vector_results": 0,
            "total_candidates_evaluated": len(curated_results),
            "total": len(results),
            "results": results,
            "feedback_session_id": feedback_session_id,
            "feedback_learning": feedback_learning,
        }

    understanding = await query_understanding_service.analyze(q, "kbji")
    retrieval_query, semantic_query, rerank_context = build_retrieval_queries(
        q,
        understanding,
    )

    bm25_top = 0
    vector_top = 0
    total_candidates = 0

    if kbji_hybrid_ready:
        retrieval = await search_kbji_hybrid_candidates(
            q,
            retrieval_top_k=20,
            retrieval_query=retrieval_query,
            semantic_query=semantic_query,
        )
        candidates = retrieval["results"]
        bm25_top = retrieval["bm25_top"]
        vector_top = retrieval["vector_top"]
        total_candidates = retrieval["total_candidates"]
        base_method = "kbji_hybrid"
    else:
        candidates = search_kbji_entries(
            retrieval_query,
            limit=15,
            original_query=q,
        )
        base_method = "kbji_local_keyword"

    if async_openai_client:
        results, reranked = await rerank_kbji_candidates(
            q,
            candidates,
            limit,
            query_context=rerank_context,
        )
        method = f"{base_method}_llm_reranked" if reranked else f"{base_method}_rerank_fallback"
        rerank_status = "success" if reranked else "fallback"
    else:
        results = candidates[:limit]
        method = base_method
        rerank_status = "disabled"

    results, feedback_session_id, feedback_learning = attach_feedback_session(
        "kbji",
        q,
        results,
        method,
        client_id=client_id,
        understanding=understanding,
    )

    return {
        "query": q,
        "method": method,
        "rerank_status": rerank_status,
        "query_understanding": understanding,
        "bm25_results": bm25_top,
        "vector_results": vector_top,
        "total_candidates_evaluated": total_candidates,
        "total": len(results),
        "results": results,
        "feedback_session_id": feedback_session_id,
        "feedback_learning": feedback_learning,
    }

@app.get("/search/kbji/status")
async def kbji_search_status():
    """Check if KBJI keyword and semantic search components are ready."""
    return {
        "status": "ready" if kbji_raw_data else "not_ready",
        "documents_indexed": len(kbji_raw_data),
        "detailed_jobs": sum(entry.get("level") == "rinci" for entry in kbji_raw_data),
        "subgroups": sum(entry.get("level") == "subgolongan" for entry in kbji_raw_data),
        "aliased_documents": sum(bool(entry.get("search_aliases")) for entry in kbji_raw_data),
        "bm25_ready": bool(kbji_bm25),
        "bm25_terms": len(kbji_bm25.idf) if kbji_bm25 else 0,
        "vector_store_ready": bool(kbji_vector_store and kbji_vector_store.is_ready),
        "hybrid_ready": kbji_hybrid_ready,
        "rerank_model": KBJI_RERANK_MODEL if async_openai_client else None,
        "query_understanding_model": QUERY_UNDERSTANDING_MODEL if async_openai_client else None,
        "embedding_model": (
            kbji_vector_store.EMBEDDING_MODEL
            if kbji_vector_store and kbji_vector_store.is_ready
            else None
        ),
    }

@app.get("/autocomplete/smart")
async def smart_autocomplete(
    q: str = Query(min_length=2, max_length=100),
    limit: int = Query(default=5, ge=1, le=20),
):
    """
    AI-Enhanced Autocomplete.
    Uses semantic understanding to provide better suggestions.
    """
    if not q or len(q) < 2:
        return {"suggestions": []}
    
    # Get AI expansion
    expansion = await expand_query_with_ai(q)
    
    # Search with expanded query
    results = search_with_keywords(expansion["keywords"], limit)
    
    suggestions = []
    for r in results:
        suggestions.append({
            "type": "smart",
            "code": r["code"],
            "judul": r["judul"],
            "match": f"{r['code']} - {r['judul'][:50]}...",
            "score": r["score"]
        })
    
    return {
        "query": q,
        "expansion": expansion.get("expanded", q),
        "suggestions": suggestions
    }

# ============================================================================
# HYBRID SEARCH ENDPOINT (NEW - v3.0)
# ============================================================================

@app.get("/search/hybrid")
async def hybrid_search(
    q: str = Query(min_length=2, max_length=200),
    top_k: int = Query(default=5, ge=1, le=10),
    use_reranking: bool = True,
    client_id: Optional[str] = Query(default=None, max_length=80),
):
    """
    🚀 Hybrid Search - Best accuracy for KBLI classification.
    
    Combines multiple retrieval methods:
    1. BM25 keyword matching (handles exact terms)
    2. Vector semantic search (handles synonyms, context)
    3. Reciprocal Rank Fusion (combines rankings)
    4. LLM semantic re-ranking (validates relevance)
    
    Args:
        q: Search query (Indonesian, can be informal)
        top_k: Number of results to return (default: 5)
        use_reranking: Whether to use LLM re-ranking (default: True)
    
    Example queries:
        - "jualan nasi goreng pinggir jalan" -> 56104, 47826
        - "tukang ojek online" -> 49422
        - "warung madura jual rokok" -> 47111
        - "konveksi baju muslim" -> 14111
    
    Returns:
        - results: List of KBLI matches with relevance scores
        - reasoning: LLM explanation for each match
    """
    if not q or len(q) < 2:
        return {
            "query": q,
            "error": "Query too short (min 2 characters)",
            "results": []
        }
    
    if not hybrid_search_engine:
        # Fallback to smart search if hybrid not available
        return await smart_search(q, limit=top_k, client_id=client_id)
    
    try:
        understanding = await query_understanding_service.analyze(q, "kbli")
        retrieval_query, semantic_query, rerank_context = build_retrieval_queries(
            q,
            understanding,
        )

        # Hybrid retrieval previously skipped the deterministic informal-term
        # expansions used by smart search. Numeric hints stay in keyword search;
        # descriptive terms enrich both sparse and semantic retrieval here.
        local_terms = [
            term for term in expand_local_keywords(q)
            if term and not str(term).isdigit()
        ]
        if local_terms:
            local_context = " ".join(dict.fromkeys(local_terms))
            retrieval_query = f"{retrieval_query} {local_context}".strip()
            semantic_query = f"{semantic_query}\n{local_context}".strip()
            rerank_context = (
                f"{rerank_context}. Istilah KBLI lokal terkait: {local_context}"
            )

        # Perform hybrid search
        result = await hybrid_search_engine.search(
            query=q,
            top_k=top_k,
            use_reranking=use_reranking,
            retrieval_query=retrieval_query,
            semantic_query=semantic_query,
            query_context=rerank_context,
        )
        
        # Format results for API response
        formatted_results = []
        for r in result.get("results", []):
            formatted = {
                "code": r.get("kode_kbli", r.get("kode", "")),
                "judul": r.get("judul", ""),
                "hierarki": r.get("hierarki", ""),
                "cakupan": r.get("cakupan", "")[:300],
                "relevance_score": r.get("relevance_score", r.get("rrf_score", 0)),
                "reasoning": r.get("reasoning", "")
            }
            if not formatted["reasoning"]:
                formatted["reasoning"] = build_local_reasoning(q, formatted)
            formatted_results.append(formatted)

        formatted_results = merge_manual_results(q, formatted_results, top_k)
        formatted_results, feedback_session_id, feedback_learning = attach_feedback_session(
            "kbli",
            q,
            formatted_results,
            "hybrid",
            client_id=client_id,
            understanding=understanding,
        )
        
        return {
            "query": q,
            "method": "hybrid",
            "query_understanding": understanding,
            "total_candidates_evaluated": result.get("total_candidates", 0),
            "bm25_results": result.get("bm25_top", 0),
            "vector_results": result.get("vector_top", 0),
            "use_reranking": use_reranking,
            "rerank_status": result.get("rerank_status", "unknown"),
            "results": formatted_results,
            "feedback_session_id": feedback_session_id,
            "feedback_learning": feedback_learning,
        }
        
    except Exception as e:
        print(f"Hybrid search error: {e}")
        # Fallback to smart search on error
        return await smart_search(q, limit=top_k, client_id=client_id)

@app.get("/search/hybrid/status")
async def hybrid_search_status():
    """
    Check if Hybrid Search Engine is available and ready.
    """
    if hybrid_search_engine and hybrid_search_engine.is_ready:
        status = {
            "status": "ready",
            "documents_indexed": len(hybrid_search_engine.documents),
            "bm25_terms": len(hybrid_search_engine.bm25.idf),
            "vector_store_ready": hybrid_search_engine.vector_store.is_ready,
            "embedding_model": hybrid_search_engine.vector_store.EMBEDDING_MODEL,
            "kbli_rerank_model": hybrid_search_engine.reranker.model,
            "query_understanding_model": QUERY_UNDERSTANDING_MODEL,
            "kbji_status": "ready" if kbji_hybrid_ready else "not_ready",
            "kbji_documents_indexed": len(kbji_raw_data) if kbji_hybrid_ready else 0,
            "kbji_bm25_terms": len(kbji_bm25.idf) if kbji_bm25 else 0,
            "kbji_vector_store_ready": bool(kbji_vector_store and kbji_vector_store.is_ready)
        }
        return status
    else:
        return {
            "status": "not_ready",
            "reason": "Hybrid Search Engine not initialized. Check OPENAI_API_KEY."
        }

@app.post("/lookup")
async def lookup_single(request: LookupRequest):
    """Lookup a single KBLI code"""
    result = lookup_code(request.code)
    return {
        "code": result["kode"],
        "found": result["status"] == "found",
        "judul": result["judul"],
        "hierarki": result["hierarki"],
        "cakupan": result["cakupan"],
        "reasoning": (
            f"Kode {result['kode']} ditemukan langsung di database KBLI."
            if result["status"] == "found"
            else "Kode tidak ditemukan di database lokal; pastikan kode 5 digit benar."
        )
    }

@app.get("/lookup/{code}")
async def lookup_code_get(code: str):
    """Lookup a single KBLI code via GET"""
    result = lookup_code(code)
    return {
        "code": result["kode"],
        "found": result["status"] == "found",
        "judul": result["judul"],
        "hierarki": result["hierarki"],
        "cakupan": result["cakupan"],
        "reasoning": (
            f"Kode {result['kode']} ditemukan langsung di database KBLI."
            if result["status"] == "found"
            else "Kode tidak ditemukan di database lokal; pastikan kode 5 digit benar."
        )
    }

@app.get("/lookup/kbji/{code}")
async def lookup_kbji_code_get(code: str):
    """Lookup a single KBJI code via GET"""
    if code in kbji_lookup:
        entry = kbji_lookup[code]
        return {
            "code": code,
            "found": True,
            "judul": entry.get("judul", ""),
            "level": entry.get("level", ""),
            "deskripsi": entry.get("deskripsi", ""),
            "reasoning": f"Kode KBJI {code} ditemukan di database KBJI."
        }
    return {
        "code": code,
        "found": False,
        "judul": "",
        "level": "",
        "deskripsi": "",
        "reasoning": "Kode KBJI tidak ditemukan di database lokal."
    }

@app.post("/upload-preview")
async def upload_preview(file: UploadFile = File(...)):
    """
    Upload Excel file and return column headers + preview.
    Does NOT process yet.
    """
    content = await read_xlsx_upload(file)
    wb = await asyncio.to_thread(load_xlsx, content, read_only=True)
    sheet = wb.active
    
    # Get headers (first row)
    headers = []
    for cell in next(sheet.iter_rows(min_row=1, max_row=1)):
        headers.append(cell.value or f"Column_{len(headers)+1}")
    
    # Get sample data (first 5 rows)
    sample_rows = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=6, values_only=True)):
        sample_rows.append(list(row))
    
    # read_only worksheets already expose the parsed worksheet bounds.
    total_rows = max(sheet.max_row - 1, 0)
    
    wb.close()
    
    return {
        "filename": file.filename,
        "headers": headers,
        "sample_data": sample_rows,
        "total_rows": total_rows
    }

@app.post("/lookup/batch")
async def lookup_batch(
    file: UploadFile = File(...),
    column_name: str = Form(...)
):
    """
    Process entire Excel file and return new Excel with lookup results.
    Pattern matching - fast and scalable.
    Returns Excel file directly.
    """
    content = await read_xlsx_upload(file)
    wb = await asyncio.to_thread(load_xlsx, content)
    sheet = wb.active
    
    # Find column index
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    try:
        col_idx = headers.index(column_name) + 1  # 1-indexed for openpyxl
    except ValueError:
        raise HTTPException(
            status_code=400,
            detail=f"Column '{column_name}' not found. Available: {headers}"
        )
    
    # Add result columns
    result_col_judul = len(headers) + 1
    result_col_hierarki = len(headers) + 2
    result_col_status = len(headers) + 3
    
    # Set headers for new columns
    sheet.cell(row=1, column=result_col_judul, value="KBLI_Judul")
    sheet.cell(row=1, column=result_col_hierarki, value="KBLI_Hierarki")
    sheet.cell(row=1, column=result_col_status, value="Lookup_Status")
    
    # Style headers
    header_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in [result_col_judul, result_col_hierarki, result_col_status]:
        cell = sheet.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
    
    # Process each row
    found_count = 0
    not_found_count = 0
    total_rows = 0
    
    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
        total_rows += 1
        cell_value = row[col_idx - 1].value
        
        if cell_value:
            # Extract KBLI codes from cell
            codes = extract_kbli_codes(str(cell_value))
            
            if codes:
                # Lookup ALL codes
                juduls, hierarkis, valid_codes = format_code_matches(codes)
                found_any = bool(valid_codes)
                
                # Join with newlines
                sheet.cell(row=row_idx, column=result_col_judul, value="\n".join(juduls))
                sheet.cell(row=row_idx, column=result_col_hierarki, value="\n".join(hierarkis))
                
                # Enable wrap text for multiline
                sheet.cell(row=row_idx, column=result_col_judul).alignment = Alignment(wrap_text=True)
                sheet.cell(row=row_idx, column=result_col_hierarki).alignment = Alignment(wrap_text=True)
                
                if found_any:
                    status_text = f"Found ({len(valid_codes)}/{len(codes)})"
                    sheet.cell(row=row_idx, column=result_col_status, value=status_text)
                    sheet.cell(row=row_idx, column=result_col_status).font = Font(color="22C55E")
                    found_count += 1
                else:
                    sheet.cell(row=row_idx, column=result_col_status, value="✗ Not Found")
                    sheet.cell(row=row_idx, column=result_col_status).font = Font(color="EF4444")
                    not_found_count += 1
            else:
                sheet.cell(row=row_idx, column=result_col_judul, value="")
                sheet.cell(row=row_idx, column=result_col_hierarki, value="")
                sheet.cell(row=row_idx, column=result_col_status, value="No code detected")
                sheet.cell(row=row_idx, column=result_col_status).font = Font(color="F59E0B")
                not_found_count += 1
        else:
            sheet.cell(row=row_idx, column=result_col_status, value="Empty cell")
            sheet.cell(row=row_idx, column=result_col_status).font = Font(color="94A3B8")
    
    # Auto-adjust column widths
    for col in [result_col_judul, result_col_hierarki, result_col_status]:
        sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 40
    
    # Save to BytesIO
    output = BytesIO()
    await asyncio.to_thread(wb.save, output)
    output.seek(0)
    wb.close()
    
    # Generate filename
    original_name = safe_output_stem(file.filename)
    result_filename = f"{original_name}_KBLI_result.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={result_filename}",
            "X-Total-Rows": str(total_rows),
            "X-Found-Count": str(found_count),
            "X-Not-Found-Count": str(not_found_count)
        }
    )


# Create temp directory for downloads
TEMP_DIR = Path(__file__).parent / "temp_downloads"
TEMP_DIR.mkdir(exist_ok=True)
TEMP_DIR_RESOLVED = TEMP_DIR.resolve()

@app.get("/download/{filename}")
async def download_file(filename: str):
    """Download generated result file"""
    if Path(filename).name != filename or not filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Invalid download filename")

    file_path = (TEMP_DIR / filename).resolve()
    if file_path.parent != TEMP_DIR_RESOLVED or not file_path.is_file():
        raise HTTPException(status_code=404, detail="File not found")
    
    return FileResponse(
        file_path, 
        filename=filename,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        background=BackgroundTask(file_path.unlink, missing_ok=True),
    )

@app.post("/lookup/batch-stream")
async def lookup_batch_stream(
    file: UploadFile = File(...),
    column_name: str = Form(...)
):
    """
    Process Excel with SSE streaming for progress updates.
    Returns progress events, then saves file and returns download URL.
    """
    content = await read_xlsx_upload(file)
    
    # Store filename for later use
    original_filename = file.filename
    wb = await asyncio.to_thread(load_xlsx, content)

    async def generate():
        sheet = wb.active
        
        # Find column index
        headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        try:
            col_idx = headers.index(column_name) + 1
        except ValueError:
            yield f"data: {json.dumps({'type': 'error', 'message': f'Column not found: {column_name}'})}\n\n"
            return
        
        # Count total rows first
        total_rows = sheet.max_row - 1
        yield f"data: {json.dumps({'type': 'start', 'total': total_rows})}\n\n"
        
        # Add result columns (same logic as before...)
        result_col_judul = len(headers) + 1
        result_col_hierarki = len(headers) + 2
        result_col_status = len(headers) + 3
        
        sheet.cell(row=1, column=result_col_judul, value="KBLI_Judul")
        sheet.cell(row=1, column=result_col_hierarki, value="KBLI_Hierarki")
        sheet.cell(row=1, column=result_col_status, value="Lookup_Status")
        
        found_count = 0
        not_found_count = 0
        
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row), start=2):
            cell_value = row[col_idx - 1].value
            current = row_idx - 1
            result_info = {"code": "", "judul": "", "status": "empty"}
            
            if cell_value:
                codes = extract_kbli_codes(str(cell_value))
                if codes:
                    juduls, hierarkis, valid_codes = format_code_matches(codes)
                    found_any = bool(valid_codes)
                    
                    if found_any:
                        sheet.cell(row=row_idx, column=result_col_judul, value="; ".join(juduls))
                        sheet.cell(row=row_idx, column=result_col_hierarki, value="; ".join(hierarkis))
                        sheet.cell(
                            row=row_idx,
                            column=result_col_status,
                            value=f"Found ({len(valid_codes)}/{len(codes)})",
                        )
                        found_count += 1
                        result_info = {"code": valid_codes[0], "judul": juduls[0], "status": "found"}
                    else:
                        sheet.cell(row=row_idx, column=result_col_status, value="Not Found")
                        not_found_count += 1
                        result_info = {"code": f"{len(codes)} codes", "judul": "", "status": "not_found"}
                else:
                     sheet.cell(row=row_idx, column=result_col_status, value="No Code")
                     not_found_count += 1
            else:
                 sheet.cell(row=row_idx, column=result_col_status, value="Empty")

            # Send progress
            if current % 10 == 0 or current == total_rows:
                yield f"data: {json.dumps({'type': 'progress', 'current': current, 'total': total_rows, 'found': found_count, 'not_found': not_found_count, 'latest': result_info})}\n\n"
        
        # Save to TEMP file instead of returning base64
        original_name_stem = safe_output_stem(original_filename)
        result_filename = f"{original_name_stem}_{uuid4().hex[:12]}_RESULT.xlsx"
        save_path = TEMP_DIR / result_filename
        
        await asyncio.to_thread(wb.save, save_path)
        wb.close()
        
        # Return download URL instead of file content
        download_url = f"/download/{result_filename}"
        
        yield f"data: {json.dumps({'type': 'complete', 'total': total_rows, 'found': found_count, 'not_found': not_found_count, 'download_url': download_url})}\n\n"

    return StreamingResponse(
        generate(),
        media_type="text/event-stream"
    )


if __name__ == "__main__":
    import uvicorn

    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=False)
